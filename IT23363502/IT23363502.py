from playwright.sync_api import sync_playwright
import openpyxl

EXCEL_FILE = "Assignment 1 - Test cases.xlsx"
URL = "https://www.pixelssuite.com/chat-translator"

def run_test():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.goto(URL)

        for row in range(2, ws.max_row + 1):
            input_text = ws.cell(row=row, column=3).value  # Input column
            expected = ws.cell(row=row, column=4).value    # Expected column

            if not input_text:
                continue

            page.fill("textarea", str(input_text))
            page.wait_for_timeout(2000)

            page.wait_for_timeout(4000)
            output = page.locator("textarea").nth(1).input_value()

            ws.cell(row=row, column=5).value = output  # Actual Output

            if expected:
                status = "PASS" if output == expected else "FAIL"
            else:
                status = "CHECK"

            ws.cell(row=row, column=6).value = status

            print(f"Row {row} -> {status}")

        wb.save(EXCEL_FILE)
        browser.close()

if __name__ == "__main__":
    run_test()